from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from os.path import exists

import pandas as pd
import matplotlib.pyplot as plt

week = {
	"Monday": {
		"Hours": 2,
	},
	"Tuesday": {
		"Hours": 6,
	},
	"Wednesday": {
		"Hours": 5,
	},
	"Thursday": {
		"Hours": 7,
	},
	"Friday": {
		"Hours": 4,
	}
}

if not exists('Overview.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 15
    ws.title = "Weekly Goal"

    headings = ['Day of the Week'] + ['Hours']
    ws.append(headings)

    for dotw in week:
        hours = list(week[dotw].values())
        ws.append([dotw] + hours)

    for col in range(1, 3):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)
    wb.save("Overview.xlsx")

workbook = 'Overview.xlsx'
df = pd.read_excel(workbook)

values = df[['Day of the Week', 'Hours']]
print(values)

ax = values.plot.bar(x='Day of the Week', y='Hours', rot=0)
plt.show()