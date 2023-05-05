from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from .models import *
from .forms import ActivityForm
from django.contrib import messages
import datetime
from django.contrib.auth.decorators import login_required
import csv
import openpyxl
from datetime import datetime, timedelta
from django.shortcuts import render
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import base64


@login_required(login_url='login')
def timetable(request):
    user = User.objects.get(username = request.user.username)
    try:
        activity_obj = Activity.objects.filter(user_id = user)
    except:
        activity_obj = None
    context = {'activity': activity_obj}
    return render(request, 'timetable.html',context)

def info(request):
    return render(request, 'info.html')

@login_required(login_url='login')
def new_activity(request):
    form = ActivityForm()
    if request.method == "POST":
        form = ActivityForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('activity_name')
            duration = form.cleaned_data.get('duration')
            description = form.cleaned_data.get('record_description')
            # end = form.cleaned_data.get('date_time_finish')
            start = form.cleaned_data.get('date_time_start')
            now = datetime.today()
            start = start.replace(tzinfo = None)
            # end = end.replace(tzinfo = None)
            user = User.objects.get(username = request.user.username)
            if now <= start:
                Activity.objects.create(
                    activity_name = name,
                    user_id = user,
                    date_time_start = start,
                    # date_time_finish = end,
                    duration = duration,
                    completion = 0,
                    record_description = description,
                    # date_time_added = now
                )
            else:
                messages.info(request, "Start date before due and start date after current date")

    context = {'form': form}
    return render(request, 'create-activity.html' , context)


@login_required(login_url='login')
def timer(request, activity_name):
    activity = Activity.objects.get(activity_name = activity_name)
    time = int((activity.duration*3600 - activity.completion*3600) // 1)
    context = {'time': time}
    return render(request, "timer.html", context)

    
def timer_stop(request):
    if request.method == 'POST':
        timer_value = request.POST.get('time')
        print(timer_value)
        # Do something with the timer value, such as saving it to a model or file
        response_data = {'message': 'Timer value received: ' + timer_value}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'message': 'Invalid request method'})



@login_required(login_url='login')
def overview(request):

    goal_value = 0.5

    # load the Excel file
    workbook = openpyxl.load_workbook('subactivities.xlsx')
    worksheet = workbook.active

    # get current date
    end_date = datetime.now().date()

    # get date one week ago
    start_date = (datetime.now() - timedelta(days=7)).date()

    # find the column indices for "name" and "date"
    name_col_index = None
    date_col_index = None
    hour_col_index = None
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        if col[0].value == "name":
            name_col_index = col[0].col_idx - 1
        elif col[0].value == "date":
            date_col_index = col[0].col_idx - 1
        elif col[0].value == "duration":
            hour_col_index = col[0].col_idx - 1
        
        if name_col_index is not None and date_col_index is not None and hour_col_index is not None:
            break

    # iterate through the rows of the worksheet to find the activities within the date range
    data = {}
    for row in worksheet.iter_rows(min_row=2):
        activity_date = row[date_col_index].value.date()
        if activity_date >= start_date and activity_date <= end_date:
            activity_label = activity_date.strftime("%d/%m")
            if not data.get(activity_label):
                data.update({activity_label: 0})
            data[activity_label] += row[hour_col_index].value

    sorted(data)

    goal = { x:goal_value for x in data}
    
    plt.plot(list(data.keys()), list(data.values()), color="blue")
    plt.plot(list(goal.keys()), list(goal.values()), color="purple")

    plt.xlabel('Day')
    plt.ylabel('Hours')

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', transparent=True)
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()

    graphic = base64.b64encode(image_png)
    graphic = graphic.decode('utf-8')


    data = list(data)
    streak = 0
    for value in data:
        if int(value[1]) >= goal_value:
            streak += 1
        else:
            streak = 0

    return render(request, 'overview.html', {'graphic':graphic, 'streak':streak})
